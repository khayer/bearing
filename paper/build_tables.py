#!/usr/bin/env python3
"""
build_tables.py -- assemble the BEARING manuscript tables workbook from the
pipeline outputs, with provenance, instead of transcribing them by hand.

WHY THIS EXISTS
---------------
Table 1 and Tables S1-S14 were previously assembled by hand from terminal output.
The NUMBERS were reproducible -- every one traces to a script -- but the WORKBOOK
was not, so a rerun of the pipeline did not update the tables and nothing
detected the drift. This script closes that gap.

It also removes a whole class of error by construction. Table S9 (parameters)
previously stated a differential testing floor of 0.1, because that is the
per-sample min_signal in config.yaml. The floor actually used was 0.5, inherited
from bearing_pvalue.py's default because the config key was unset. A transcribed
table cannot catch that; a generated one can. Table S9 is therefore DERIVED here
from config.yaml plus the defaults parsed out of the scripts themselves, and
build_tables.py fails loudly if it cannot resolve them.

DESIGN RULES
------------
1. NEVER fabricate. If a source file is missing, the sheet is not written and the
   run fails (or, with --allow-missing, the sheet is written with an explicit
   MISSING-SOURCE banner and the manifest records it). A blank cell must never be
   mistakable for a measured zero.
2. Every sheet records its provenance: source path, SHA256, mtime, row count.
   The Provenance sheet is part of the workbook, not a side file.
3. Staleness is checked, not assumed. Any source older than the p-value layer
   (results/pvalue.done) is flagged -- that is the same rule as staleness_audit.sh.
4. Curated content (cell-type metadata, the tool-capability comparison) lives in
   a version-controlled YAML, not in this script and not in someone's memory.

USAGE
-----
    # what would be built, and what is missing:
    python3 build_tables.py --results-dir workflow/results --dry-run

    # build it:
    python3 build_tables.py \
        --results-dir workflow/results \
        --config workflow/config/config.yaml \
        --sheet workflow/config/samples.tsv \
        --curated paper/tables_curated.yaml \
        --repo . \
        --out BEARING_tables.xlsx

ASCII only.
"""

import argparse
import csv
import datetime as _dt
import hashlib
import os
import re
import sys

try:
    import yaml
except ImportError:
    yaml = None

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


HDR_FILL = PatternFill("solid", fgColor="DDDDDD")
WARN_FILL = PatternFill("solid", fgColor="FFC7CE")
GEN_FILL = PatternFill("solid", fgColor="E2EFDA")
BOLD = Font(bold=True)


# ---------------------------------------------------------------------------
# provenance
# ---------------------------------------------------------------------------

class Provenance(object):
    """Records every file this build read, so a table can be traced or falsified."""

    def __init__(self, reference=None):
        self.rows = []
        self.reference_mtime = None
        self.reference_path = reference
        if reference and os.path.exists(reference):
            self.reference_mtime = os.path.getmtime(reference)

    def record(self, sheet, path, n_rows=None, note=""):
        exists = os.path.exists(path)
        sha = mtime = ""
        stale = ""
        if exists:
            h = hashlib.sha256()
            with open(path, "rb") as fh:
                for chunk in iter(lambda: fh.read(1 << 20), b""):
                    h.update(chunk)
            sha = h.hexdigest()[:16]
            mt = os.path.getmtime(path)
            mtime = _dt.datetime.fromtimestamp(mt).strftime("%Y-%m-%d %H:%M:%S")
            if self.reference_mtime and mt < self.reference_mtime:
                stale = "STALE (older than %s)" % os.path.basename(self.reference_path)
        self.rows.append({
            "sheet": sheet,
            "source": path,
            "exists": "yes" if exists else "NO",
            "sha256_16": sha,
            "mtime": mtime,
            "rows": "" if n_rows is None else n_rows,
            "stale": stale,
            "note": note,
        })
        return exists


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def read_tsv(path):
    """Return (header, rows). Tolerates '#' comment lines, keeps them separate."""
    header, rows, comments = None, [], []
    with open(path, newline="") as fh:
        for line in fh:
            line = line.rstrip("\n")
            if not line:
                continue
            if line.startswith("#"):
                comments.append(line)
                continue
            f = line.split("\t")
            if header is None:
                header = f
            else:
                rows.append(f)
    return header or [], rows, comments


def write_grid(ws, header, rows, start=1, comments=None):
    r = start
    if comments:
        for c in comments:
            ws.cell(r, 1, c).font = Font(italic=True, color="666666")
            r += 1
        r += 1
    for j, h in enumerate(header, 1):
        c = ws.cell(r, j, h)
        c.font = BOLD
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
    for row in rows:
        for j, v in enumerate(row, 1):
            try:
                ws.cell(r, j, float(v) if re.match(r"^-?\d+\.?\d*([eE][-+]?\d+)?$", v) else v)
            except (ValueError, TypeError):
                ws.cell(r, j, v)
        r += 1
    return r


def banner_missing(ws, path):
    ws.cell(1, 1, "MISSING SOURCE -- THIS SHEET WAS NOT BUILT").font = Font(bold=True, color="9C0006")
    ws.cell(1, 1).fill = WARN_FILL
    ws.cell(2, 1, "Expected: %s" % path)
    ws.cell(3, 1, "Run the producing step (see paper/reproduce_all.sh), then rebuild.")
    ws.cell(4, 1, "No values are shown because none were read. Do not interpret this "
                  "sheet as zeros or as absence of an effect.")


# ---------------------------------------------------------------------------
# Table S9 -- DERIVED, never transcribed
# ---------------------------------------------------------------------------

def resolve_pvalue_floor(repo, config, score_method="kl"):
    """
    The differential testing floor ACTUALLY in force.

    If config sets pvalue_min_signal, that wins. Otherwise bearing_pvalue.py
    resolves its own default by score method -- parse it out of the source so
    that a change to the default cannot silently invalidate the table.
    """
    cfg = config.get("pvalue_min_signal", None)
    if cfg is not None:
        return float(cfg), "config.yaml: pvalue_min_signal"

    src_path = os.path.join(repo, "bearing_pvalue.py")
    if not os.path.exists(src_path):
        sys.exit("ERROR: cannot find %s to resolve the differential floor default. "
                 "Pass --repo." % src_path)
    src = open(src_path).read()
    m = re.search(
        r"args\.min_signal\s*=\s*([0-9.]+)\s+if\s+args\.score_method\s*==\s*[\"']kl[\"']"
        r"\s+else\s+([0-9.]+)", src)
    if not m:
        sys.exit("ERROR: could not parse the --min-signal default out of "
                 "bearing_pvalue.py. The code changed; fix resolve_pvalue_floor() "
                 "rather than hard-coding a number that may now be wrong.")
    val = float(m.group(1) if score_method == "kl" else m.group(2))
    return val, ("bearing_pvalue.py default for score_method=%s "
                 "(config key pvalue_min_signal is UNSET)" % score_method)


def resolve_qcat_floor(repo, config):
    """
    Per-sample raw-signal floor: the config value, plus the script default for
    context.

    bigwig_to_qcat.py declares `default=MIN_SIGNAL`, a module constant, so a
    regex for a numeric literal finds nothing and would silently report the
    default as unknown. Resolve the literal first, then the constant. If neither
    can be resolved, say so rather than printing a plausible-looking wrong value.
    """
    cfg = config.get("min_signal", None)
    src_path = os.path.join(repo, "bigwig_to_qcat.py")
    if not os.path.exists(src_path):
        return cfg, None
    src = open(src_path).read()

    m = re.search(r"[\"']--min-signal[\"'][^)]*?default\s*=\s*([A-Za-z_0-9.]+)", src, re.S)
    if not m:
        return cfg, None
    tok = m.group(1)
    if re.match(r"^[0-9.]+$", tok):
        return cfg, float(tok)
    # default is a module constant -- resolve its assignment
    m2 = re.search(r"^%s\s*=\s*([0-9.]+)" % re.escape(tok), src, re.M)
    if m2:
        return cfg, float(m2.group(1))
    return cfg, None


def build_s9(ws, config, repo, prov):
    ws.cell(1, 1, "Table S9. Parameters").font = Font(bold=True, size=12)
    ws.cell(2, 1, "GENERATED from config.yaml and from the defaults parsed out of the "
                  "scripts. Do not edit by hand: edit config.yaml and rebuild.").font = \
        Font(italic=True, color="666666")

    pv_floor, pv_src = resolve_pvalue_floor(repo, config)
    qc_cfg, qc_default = resolve_qcat_floor(repo, config)

    rows = [
        ["min_signal (per-sample raw signal)", qc_cfg,
         ("bigwig_to_qcat.py --min-signal (script default %s)" % qc_default)
         if qc_default is not None else
         "config.yaml: min_signal (script default UNRESOLVED - see build_tables.py)",
         "Bins whose total raw signal falls below this have all per-track scores "
         "set to zero. Calibration: Methods M.7."],
        ["Differential testing floor (|BES|)", pv_floor, pv_src,
         "Bins whose |summed differential| falls below this are NOT TESTED and are "
         "absent from the differential stats table. DISTINCT from min_signal above. "
         "Sensitivity: Table S14, Supplementary Figure S11."],
        ["n_perms", config.get("n_perms"), "config.yaml: n_perms", ""],
        ["min_shift", config.get("min_shift"), "config.yaml: min_shift", ""],
        ["seed", config.get("seed"), "config.yaml: seed", ""],
        ["normalize", config.get("normalize") or "(off)", "config.yaml: normalize",
         "Empty = no normalization (Methods M.2)."],
        ["reference_condition", config.get("reference_condition"),
         "config.yaml: reference_condition", ""],
        ["blacklist_merge_distance", config.get("blacklist_merge_distance"),
         "config.yaml: blacklist_merge_distance", ""],
        ["GTF", os.path.basename(str(config.get("gtf", ""))), "config.yaml: gtf", ""],
    ]
    hdr = ["Parameter", "Value", "Source", "Notes"]
    r = write_grid(ws, hdr, [[str(x) if x is not None else "" for x in row] for row in rows], start=4)

    # flag the trap that caused the original error
    if config.get("pvalue_min_signal") is None:
        c = ws.cell(r + 1, 1,
                    "WARNING: pvalue_min_signal is not set in config.yaml, so the "
                    "differential floor above comes from a library default and will "
                    "change silently if that default changes. Set it explicitly.")
        c.font = Font(bold=True, color="9C0006")
        c.fill = WARN_FILL
    for col, w in zip("ABCD", [34, 16, 40, 72]):
        ws.column_dimensions[col].width = w
    prov.record("Table S9 - parameters", os.path.join(repo, "bearing_pvalue.py"),
                note="parsed for the --min-signal default")
    return True


# ---------------------------------------------------------------------------
# Table S4 -- DERIVED from the sample sheet
# ---------------------------------------------------------------------------

def build_s4(ws, sheet_path, prov):
    if not prov.record("Table S4 - accessions", sheet_path):
        banner_missing(ws, sheet_path)
        return False
    hdr, rows, _ = read_tsv(sheet_path)
    ws.cell(1, 1, "Table S4. Samples and accessions").font = Font(bold=True, size=12)
    ws.cell(2, 1, "GENERATED from the sample sheet.").font = Font(italic=True, color="666666")
    write_grid(ws, hdr, rows, start=4)
    for j in range(1, len(hdr) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 22
    return True


# ---------------------------------------------------------------------------
# registry
# ---------------------------------------------------------------------------

def build_curated(ws, spec, curated_path, sheet_label):
    """
    Render a curated table (Table 1, Table S5) from tables_curated.yaml.

    Curated tables are metadata and literature, not measurements: they cannot be
    derived from a pipeline output. Keeping them in a version-controlled YAML
    means they are reviewable in a diff and assembled by script, rather than
    retyped into a spreadsheet each revision.
    """
    r = 1
    ws.cell(r, 1, spec.get("title", sheet_label)).font = Font(bold=True, size=12)
    r += 1
    if spec.get("subtitle"):
        c = ws.cell(r, 1, spec["subtitle"])
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    c = ws.cell(r, 1, "CURATED content from %s (version-controlled). Edit that file, "
                      "not this sheet." % curated_path)
    c.font = Font(italic=True, color="666666")
    c.fill = GEN_FILL
    r += 2

    r = write_grid(ws, spec["header"], [[str(x) for x in row] for row in spec["rows"]], start=r)

    if spec.get("notes"):
        r += 1
        ws.cell(r, 1, "Notes").font = BOLD
        r += 1
        for n in spec["notes"]:
            c = ws.cell(r, 1, n)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=1,
                           end_row=r, end_column=max(2, len(spec["header"])))
            ws.row_dimensions[r].height = 42
            r += 1

    # Table S5 carries a second block (related tools) with its own header
    if spec.get("related_rows"):
        r += 1
        ws.cell(r, 1, "Related differential and information-theoretic methods").font = BOLD
        r += 1
        r = write_grid(ws, spec["related_header"],
                       [[str(x) for x in row] for row in spec["related_rows"]], start=r)

    widths = [26, 30, 30, 30, 30, 22, 8, 44, 30]
    for j in range(1, len(spec["header"]) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = \
            widths[j - 1] if j <= len(widths) else 24
    return r


class Spec(object):
    """
    One sheet. `sources` is a list of glob patterns: several tables are built
    from more than one file (Table S13 has one replicate_stability_*.tsv per
    comparison), and a single hard-coded path would silently drop the rest.
    When several files match, they are concatenated and a `source_file` column
    records which file each row came from.
    """

    def __init__(self, sheet, sources, title=None, note="", expect="one",
                 label_col=None):
        # sources: list of glob strings, or (glob, label) pairs. A label is how a
        # merged table keeps a meaningful column ("min.count=10") instead of a
        # filename.
        norm = []
        for item in (sources if isinstance(sources, (list, tuple)) else [sources]):
            if isinstance(item, (list, tuple)):
                norm.append((item[0], item[1] if len(item) > 1 else None))
            else:
                norm.append((item, None))
        self.sheet = sheet
        self.sources = norm
        self.label_col = label_col
        self.title = title or sheet
        self.note = note
        # "one"  -> exactly one file should match; several means the pattern is
        #           ambiguous and we must NOT silently pick or merge.
        # "many" -> the table is genuinely assembled from several files.
        self.expect = expect

    def resolve(self):
        """Glob every pattern; de-duplicate by REAL PATH.

        Searching both `--sources-dir .` and the repo root yields './x.tsv' and
        'x.tsv' for the same file. String de-duplication keeps both and the
        sheet gets every row twice. Resolve to a canonical path first.
        """
        import glob as _g
        hits = []
        for pat, _lab in self.sources:
            hits.extend(sorted(_g.glob(pat)))
        seen, out = set(), []
        for h in hits:
            key = os.path.realpath(h)
            if key not in seen:
                seen.add(key)
                out.append(h)
        return out

    def label_for(self, path):
        """Semantic label for a file if the registry gave one, else its name."""
        import glob as _g
        rp = os.path.realpath(path)
        for pat, label in self.sources:
            if label and any(os.path.realpath(h) == rp for h in _g.glob(pat)):
                return label
        return os.path.basename(path)


def table_registry(results, sources):
    """
    Declarative map of sheet -> source glob(s).

    Two roots, because the pipeline and the hand-run analyses write to different
    places:
      results  -- Snakemake rule outputs (--results-dir)
      sources  -- outputs of the dev/ scripts, which are NOT workflow rules and
                  are run by hand (--sources-dir, default paper/table_sources).
                  These are small and ARE tracked in git, so that a fresh clone
                  can rebuild the workbook without re-running the pipeline.
                  workflow/results/ is gitignored and cannot serve that purpose.

    EACH TABLE HAS EXACTLY ONE CANONICAL LOCATION. Do not search both roots for
    the same table: two copies of one table are different files with identical
    content, so realpath de-duplication does not catch them, and a spec with
    expect="many" would concatenate them into doubled rows.

    paper/reproduce_all.sh must write the dev/ outputs to --sources-dir. If it
    writes them elsewhere the loop is broken: the scripts run, and this script
    still reports MISSING.
    """
    j = lambda *p: os.path.join(results, *p)
    s = lambda *p: os.path.join(sources, *p)
    root = lambda *p: os.path.join(*p)

    return [
        Spec("Table S1 - pairwise JSD",
             [j("compare", "*_q_pair_jsd.tsv")],
             note="compare_qcat.py write_q_pair_jsd_tsv"),
        Spec("Table S2 - regional q-values",
             [j("regional", "consolidated_enrichment_tcrb.tsv"),
              j("regional", "consolidated_enrichment_igh.tsv")],
             expect="many", note="rule regional_consolidate (one file per locus)"),
        Spec("Table S3 - CBE null",
             [j("regional", "enrich_cbe_*.tsv")],
             note="regional_enrichment.py --region-assign overlap on cbe_mm10.bed"),
        Spec("Table S6 - FDR calibration",
             [j("calibration", "calibration_summary.tsv")],
             note="rule calibration"),
        Spec("Table S7 - recovery sweep",
             [j("benchmark", "*recovery*.tsv")],
             note="benchmark rule"),
        # EXPLICIT, not a glob. Table S8 reports DN-vs-DP ONLY, at two expression
        # filters (min.count=10 and min.count=3) -- that is what its "Expression
        # filter" column distinguishes. A rna_concordance_*_summary.tsv glob also
        # matches DN_vs_EbKO_rc, DN_vs_ProB_igh, DN_vs_ProB_mc3_igh and
        # fwd_DN_vs_DP: DIFFERENT analyses, belonging to other tables or to none.
        # Merging them would produce a table that looks authoritative and is not.
        Spec("Table S8 - RNA edgeR validation",
             [(s("rna_concordance_DN_vs_DP_summary.tsv"), "min.count=10"),
              (s("rna_concordance_DN_vs_DP_mc3_summary.tsv"), "min.count=3")],
             expect="many", label_col="Expression filter",
             note="rna_concordance_stranded.R; DN-vs-DP at two expression filters"),
        Spec("Table S10 - baselines",
             [s("baseline_comparison*.tsv")],
             note="dev/baseline_comparison.py (NOT a workflow rule)"),
        # EXPLICIT, not a glob: several regional_null_calibration_* files exist
        # (tcrb/igh x within/background x different --n-random). They are
        # DIFFERENT ANALYSES and must not be concatenated. Table S11's own note
        # records the published run as background mode, n=500, seed=42 -- i.e.
        # the *_bg files. Confirm before trusting this.
        Spec("Table S11 - regional null",
             [s("regional_null_calibration_tcrb_DNrep_bg.tsv"),
              s("regional_null_calibration_igh_DNrep_bg.tsv")],
             expect="many",
             note="dev/regional_null_calibration.py --mode background --n-random 500 "
                  "--seed 42. See the ref_bins matching issue in reproduce_all.sh."),
        Spec("Table S12 - track ablation",
             [s("track_ablation_*.tsv")],
             expect="many", note="dev/track_ablation.py (NOT a workflow rule)"),
        Spec("Table S13 - replicate stability",
             [s("replicate_stability_DN_vs_*.tsv")],
             expect="many", note="dev/replicate_stability.py (NOT a workflow rule)"),
        Spec("Table S14 - floor sensitivity",
             [s("sens", "floor_sweep.tsv")],
             note="pvminsig_sweep.py (NOT a workflow rule)"),
        Spec("data - significant bins",
             [j("regional", "significant_bins_summary.tsv")],
             note="rule significant_bins_summary"),
    ]


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--results-dir", default="workflow/results")
    ap.add_argument("--config", default="workflow/config/config.yaml")
    ap.add_argument("--sheet", default="workflow/config/samples.tsv")
    ap.add_argument("--curated", default="paper/tables_curated.yaml",
                    help="Version-controlled curated content (Table 1, S5).")
    ap.add_argument("--sources-dir", default="paper/table_sources",
                    help="Outputs of the dev/ scripts, which are run by hand and "
                         "are not workflow rules. The repo root is also searched.")
    ap.add_argument("--repo", default=".", help="Repo root, for parsing script defaults.")
    ap.add_argument("--out", default="BEARING_tables.xlsx")
    ap.add_argument("--reference", default=None,
                    help="Staleness reference (default: <results>/pvalue.done).")
    ap.add_argument("--dry-run", action="store_true",
                    help="Report which sources exist; write nothing.")
    ap.add_argument("--allow-missing", action="store_true",
                    help="Write a MISSING-SOURCE banner sheet instead of failing. "
                         "The manifest records every gap.")
    a = ap.parse_args()

    ref = a.reference or os.path.join(a.results_dir, "pvalue.done")
    prov = Provenance(ref)
    if prov.reference_mtime is None:
        print("NOTE: staleness reference %s not found; staleness will not be "
              "checked." % ref, file=sys.stderr)

    specs = table_registry(a.results_dir, a.sources_dir)

    if a.dry_run:
        print("%-42s %-9s %s" % ("SHEET", "FILES", "RESOLVED / SEARCHED"))
        missing = 0
        for sp in specs:
            hits = sp.resolve()
            if not hits:
                missing += 1
                print("%-42s %-9s %s" % (sp.sheet, "MISSING", " | ".join(x[0] for x in sp.sources)))
            elif sp.expect == "one" and len(hits) > 1:
                missing += 1
                print("%-42s %-9s %s" % (sp.sheet, "AMBIGUOUS",
                                         "%d files match a pattern expecting ONE:" % len(hits)))
                for h in hits:
                    print("%-42s %-9s %s" % ("", "", h))
            else:
                print("%-42s %-9s %s" % (sp.sheet, "%d file(s)" % len(hits), hits[0]))
                for h in hits[1:]:
                    print("%-42s %-9s %s" % ("", "", h))
        print("\nderived sheets (no TSV source): Table S9 (from config.yaml), "
              "Table S4 (from the sample sheet)")
        print("curated sheets (from %s): Table 1, Table S5" % a.curated)
        print("\n%d of %d TSV sources missing." % (missing, len(specs)))
        if missing:
            print("Run the producing steps in paper/reproduce_all.sh, or fix the "
                  "paths in table_registry().")
        return 1 if missing else 0

    if yaml is None:
        sys.exit("pyyaml is required to read the config: pip install pyyaml")
    if not os.path.exists(a.config):
        sys.exit("ERROR: config not found: %s" % a.config)
    config = yaml.safe_load(open(a.config)) or {}

    curated = {}
    if os.path.exists(a.curated):
        curated = yaml.safe_load(open(a.curated)) or {}
        prov.record("(curated)", a.curated, note="Table 1, Table S5 content")
    else:
        print("WARNING: curated file %s not found; Table 1 and Table S5 will be "
              "skipped." % a.curated, file=sys.stderr)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    failures = []

    # --- curated: Table 1 -------------------------------------------------
    if "table1" in curated:
        build_curated(wb.create_sheet("Table 1 - cell types"),
                      curated["table1"], a.curated, "Table 1")

    # --- derived: S4, S9 ---------------------------------------------------
    ws = wb.create_sheet("Table S4 - accessions")
    if not build_s4(ws, a.sheet, prov) and not a.allow_missing:
        failures.append("Table S4 (%s)" % a.sheet)

    ws = wb.create_sheet("Table S9 - parameters")
    build_s9(ws, config, a.repo, prov)

    # --- curated: S5 -------------------------------------------------------
    if "table_s5" in curated:
        build_curated(wb.create_sheet("Table S5 - capability compare"),
                      curated["table_s5"], a.curated, "Table S5")

    # --- registry-driven ---------------------------------------------------
    for sp in specs:
        ws = wb.create_sheet(sp.sheet[:31])
        hits = sp.resolve()
        if not hits:
            prov.record(sp.sheet, " | ".join(x[0] for x in sp.sources), note=sp.note + " [NOT FOUND]")
            banner_missing(ws, " | ".join(x[0] for x in sp.sources))
            failures.append("%s (%s)" % (sp.sheet, sp.sources[0][0]))
            continue
        if sp.expect == "one" and len(hits) > 1:
            # Ambiguity is not resolvable by guessing: picking one, or merging
            # them, both risk silently publishing the wrong run.
            prov.record(sp.sheet, " | ".join(hits),
                        note=sp.note + " [AMBIGUOUS: %d files matched]" % len(hits))
            ws.cell(1, 1, "AMBIGUOUS SOURCE -- THIS SHEET WAS NOT BUILT").font = \
                Font(bold=True, color="9C0006")
            ws.cell(1, 1).fill = WARN_FILL
            ws.cell(2, 1, "%d files matched a pattern that expects exactly one:" % len(hits))
            for i, h in enumerate(hits):
                ws.cell(3 + i, 1, h)
            ws.cell(3 + len(hits), 1,
                    "These may be different analyses. Name the intended file "
                    "explicitly in table_registry() rather than letting a glob choose.")
            failures.append("%s (ambiguous: %d files)" % (sp.sheet, len(hits)))
            continue

        # concatenate; a source_file column keeps every row traceable
        all_hdr, all_rows, all_comments = None, [], []
        for h in hits:
            hdr, rows, comments = read_tsv(h)
            prov.record(sp.sheet, h, n_rows=len(rows), note=sp.note)
            all_comments.extend(comments)
            tag = sp.label_for(h)
            if all_hdr is None:
                lab = sp.label_col or "source_file"
                all_hdr = list(hdr) + ([lab] if len(hits) > 1 else [])
            elif hdr != all_hdr[:len(hdr)]:
                # do not silently glue together tables with different columns
                prov.rows[-1]["note"] += " [HEADER MISMATCH -- not merged]"
                failures.append("%s (header mismatch in %s)" % (sp.sheet, tag))
                continue
            for row in rows:
                all_rows.append(list(row) + ([tag] if len(hits) > 1 else []))

        if all_hdr is None:
            banner_missing(ws, " | ".join(x[0] for x in sp.sources))
            continue
        ws.cell(1, 1, sp.title).font = Font(bold=True, size=12)
        src_note = ("GENERATED from %d file(s): %s"
                    % (len(hits), ", ".join(os.path.basename(h) for h in hits)))
        c = ws.cell(2, 1, src_note)
        c.font = Font(italic=True, color="666666")
        c.fill = GEN_FILL
        write_grid(ws, all_hdr, all_rows, start=4, comments=all_comments)
        for j in range(1, min(len(all_hdr), 30) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 18

    # --- provenance sheet --------------------------------------------------
    ws = wb.create_sheet("Provenance")
    ws.cell(1, 1, "Provenance").font = Font(bold=True, size=12)
    ws.cell(2, 1, "Built %s by build_tables.py. Every sheet above traces to a source "
                  "listed here." % _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(3, 1, "Staleness reference: %s" % ref)
    hdr = ["sheet", "source", "exists", "sha256_16", "mtime", "rows", "stale", "note"]
    r = write_grid(ws, hdr, [[str(x[k]) for k in hdr] for x in prov.rows], start=5)
    for row_i, rec in enumerate(prov.rows, start=6):
        if rec["stale"] or rec["exists"] == "NO":
            for j in range(1, len(hdr) + 1):
                ws.cell(row_i, j).fill = WARN_FILL
    for col, w in zip("ABCDEFGH", [34, 56, 8, 18, 20, 8, 30, 46]):
        ws.column_dimensions[col].width = w

    stale = [x for x in prov.rows if x["stale"]]
    if stale:
        print("\nSTALE SOURCES (older than %s):" % os.path.basename(ref), file=sys.stderr)
        for x in stale:
            print("   %-40s %s" % (x["sheet"], x["source"]), file=sys.stderr)
        print("These tables were built from outputs predating the current p-value "
              "layer. Re-run the producing steps.", file=sys.stderr)

    if failures and not a.allow_missing:
        print("\nERROR: %d sheet(s) had no source and were NOT built:" % len(failures),
              file=sys.stderr)
        for f in failures:
            print("   %s" % f, file=sys.stderr)
        print("\nRefusing to write a workbook with silently empty tables. Re-run the "
              "producing steps, or pass --allow-missing to emit banner sheets that "
              "say so explicitly.", file=sys.stderr)
        return 2

    wb.save(a.out)
    print("wrote %s (%d sheets)" % (a.out, len(wb.sheetnames)))
    if failures:
        print("  %d sheet(s) carry a MISSING-SOURCE banner: %s"
              % (len(failures), ", ".join(failures)))
    if stale:
        print("  %d source(s) are STALE -- see the Provenance sheet." % len(stale))
    return 0


if __name__ == "__main__":
    sys.exit(main())
