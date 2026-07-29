#!/usr/bin/env python3
"""
Build the BEARING supplementary compartment table (dcHiC over the Tcrb window)
from the per-contrast region-extract TSVs written by dchic_region_result.py.

Inputs (in --tsv-dir, default paper/table_sources/):
  Full-panel 5-condition run (DN,DP,EbKO,ProB,3T3), global padj:
    Tcrb_compartment_dchic_S3T3_100kb.tsv   (gives DN + 3T3)
    Tcrb_compartment_dchic_DP_100kb.tsv     (gives DP)
    Tcrb_compartment_dchic_EBKO_100kb.tsv   (gives EbKO)
    Tcrb_compartment_dchic_ProB_100kb.tsv   (gives ProB)
  Lymphoid-only 4-condition run (DN,DP,EbKO,ProB), S3T3-free padj:
    Tcrb_compartment_lymphoidonly_{DP,EBKO,ProB}_100kb.tsv

Output: an .xlsx with three sheets (compartment_calls, lymphoid_only_test, notes).

Usage:
  python3 build_compartment_supp_table.py --tsv-dir paper/table_sources \
      --out paper/tables/Supp_Table_compartment_dcHiC.xlsx

ASCII-only source. No formulas (pure data table), so no recalculation needed.
"""
import argparse
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def read_tsv(path):
    """Return (header_list, list_of_row_dicts). Skips comment lines."""
    rows = []
    header = None
    with open(path) as fh:
        for line in fh:
            if line.startswith("#"):
                continue
            parts = line.rstrip("\n").split("\t")
            if header is None:
                header = parts
                continue
            rows.append(dict(zip(header, parts)))
    return header, rows


def pc_col(header, exp):
    for h in header:
        if h == exp + "_PC":
            return h
    raise SystemExit("no %s_PC column in %s" % (exp, header))


def padj_col(header):
    for name in ("padj_global", "padj_lymphoid", "padj"):
        if name in header:
            return name
    raise SystemExit("no padj column in %s" % header)


def load_contrast(path, exp_b):
    header, rows = read_tsv(path)
    pca = pc_col(header, "DN")
    pcb = pc_col(header, exp_b)
    pj = padj_col(header)
    out = {}
    for r in rows:
        key = (r["chr"], int(r["start"]), int(r["end"]))
        out[key] = {"DN": float(r[pca]), exp_b: float(r[pcb]),
                    "padj": float(r[pj])}
    return out


def call(pc):
    return "A" if pc > 0 else "B"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tsv-dir", default="paper/table_sources",
                    help="dir with the 5-condition run TSVs "
                         "(Tcrb_compartment_dchic_{S3T3,DP,EBKO,ProB}_<res>.tsv)")
    ap.add_argument("--lymphoid-dir", default=None,
                    help="dir with the lymphoid-only run TSVs "
                         "(Tcrb_compartment_dchic_{DP,EBKO,ProB}_<res>.tsv); "
                         "default: <tsv-dir>/lymphoid")
    ap.add_argument("--res", default="100kb", help="resolution label in filenames")
    ap.add_argument("--out", default="Supp_Table_compartment_dcHiC.xlsx")
    a = ap.parse_args()

    d = a.tsv_dir
    ld = a.lymphoid_dir or os.path.join(d, "lymphoid")
    R = a.res
    fp = {
        "S3T3": os.path.join(d, "Tcrb_compartment_dchic_S3T3_%s.tsv" % R),
        "DP": os.path.join(d, "Tcrb_compartment_dchic_DP_%s.tsv" % R),
        "EBKO": os.path.join(d, "Tcrb_compartment_dchic_EBKO_%s.tsv" % R),
        "ProB": os.path.join(d, "Tcrb_compartment_dchic_ProB_%s.tsv" % R),
    }
    ly = {
        "DP": os.path.join(ld, "Tcrb_compartment_dchic_DP_%s.tsv" % R),
        "EBKO": os.path.join(ld, "Tcrb_compartment_dchic_EBKO_%s.tsv" % R),
        "ProB": os.path.join(ld, "Tcrb_compartment_dchic_ProB_%s.tsv" % R),
    }
    for p in list(fp.values()) + list(ly.values()):
        if not os.path.exists(p):
            sys.exit("missing input TSV: %s" % p)

    s3 = load_contrast(fp["S3T3"], "S3T3")
    dp = load_contrast(fp["DP"], "DP")
    eb = load_contrast(fp["EBKO"], "EBKO")
    pb = load_contrast(fp["ProB"], "ProB")
    bins = sorted(s3.keys(), key=lambda k: k[1])

    ldp = load_contrast(ly["DP"], "DP")
    leb = load_contrast(ly["EBKO"], "EBKO")
    lpb = load_contrast(ly["ProB"], "ProB")

    wb = Workbook()
    ARIAL = "Arial"
    hdr_font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="305496")
    cell_font = Font(name=ARIAL, size=10)
    A_fill = PatternFill("solid", fgColor="D6E4C6")   # muted green
    B_fill = PatternFill("solid", fgColor="F2C9C9")   # muted red
    flip_font = Font(name=ARIAL, size=10, bold=True)
    center = Alignment(horizontal="center")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, ncols, row=1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = border

    # ---------- Sheet 1: compartment calls (5-condition full panel) ----------
    ws = wb.active
    ws.title = "compartment_calls"
    conds = ["DN", "DP", "EbKO", "ProB", "3T3"]
    cond_src = {"DN": None, "DP": dp, "EbKO": eb, "ProB": pb, "3T3": s3}
    cond_key = {"DP": "DP", "EbKO": "EBKO", "ProB": "ProB", "3T3": "S3T3"}
    headers = ["chr", "start", "end"]
    for cnd in conds:
        headers += [cnd + " PC1", cnd + " call"]
    headers += ["padj (5-cond, global)"]
    ws.append(headers)
    style_header(ws, len(headers))

    for key in bins:
        chrom, st, en = key
        dn_pc = s3[key]["DN"]
        vals = {"DN": dn_pc, "DP": dp[key]["DP"], "EbKO": eb[key]["EBKO"],
                "ProB": pb[key]["ProB"], "3T3": s3[key]["S3T3"]}
        row = [chrom, st, en]
        for cnd in conds:
            row += [round(vals[cnd], 4), call(vals[cnd])]
        row += [s3[key]["padj"]]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).font = cell_font
            ws.cell(row=r, column=c).border = border
        # color the call cells; bold + fill flips relative to DN
        for i, cnd in enumerate(conds):
            call_col = 3 + 2 * i + 2   # position of "<cond> call"
            cell = ws.cell(row=r, column=call_col)
            cell.alignment = center
            if vals[cnd] > 0:
                cell.fill = A_fill
            else:
                cell.fill = B_fill
            if (vals[cnd] > 0) != (dn_pc > 0) and cnd != "DN":
                cell.font = flip_font

    ws.freeze_panes = "D2"
    widths = [6, 10, 10] + [9, 7] * len(conds) + [20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    # ---------- Sheet 2: lymphoid-only test (S3T3-free padj) ----------
    ws2 = wb.create_sheet("lymphoid_only_test")
    h2 = ["chr", "start", "end", "DN PC1", "DP PC1", "EbKO PC1", "ProB PC1",
          "padj (4-cond lymphoid)", "DP flips", "EbKO flips", "ProB flips"]
    ws2.append(h2)
    style_header(ws2, len(h2))
    for key in bins:
        chrom, st, en = key
        dn = ldp[key]["DN"]
        vdp, veb, vpb = ldp[key]["DP"], leb[key]["EBKO"], lpb[key]["ProB"]
        pj = ldp[key]["padj"]
        row = [chrom, st, en, round(dn, 4), round(vdp, 4), round(veb, 4),
               round(vpb, 4), pj,
               "yes" if (vdp > 0) != (dn > 0) else "",
               "yes" if (veb > 0) != (dn > 0) else "",
               "yes" if (vpb > 0) != (dn > 0) else ""]
        ws2.append(row)
        r = ws2.max_row
        for c in range(1, len(h2) + 1):
            ws2.cell(row=r, column=c).font = cell_font
            ws2.cell(row=r, column=c).border = border
            if c >= 9:
                ws2.cell(row=r, column=c).alignment = center
        if pj <= 0.05:
            for c in range(1, len(h2) + 1):
                ws2.cell(row=r, column=c).font = flip_font
    ws2.freeze_panes = "D2"
    for col, w in zip("ABCDEFGHIJK",
                      [6, 10, 10, 9, 9, 9, 9, 20, 9, 10, 10]):
        ws2.column_dimensions[col].width = w

    # ---------- Sheet 3: notes ----------
    ws3 = wb.create_sheet("notes")
    notes = [
        ("Supplementary Table. Compartment (dcHiC) analysis of the Tcrb "
         "recombination domain.", True),
        ("", False),
        ("Window: chr6:40,400,000-42,400,000 (mm10), 100 kb bins (20 bins).",
         False),
        ("Method: dcHiC differential compartment analysis. PC1 computed by "
         "dcHiC (MFA + quantile normalization + Mahalanobis), oriented so "
         "positive PC1 = active A compartment. Control condition: DN.", False),
        ("Matrices: per-replicate KR-balanced Hi-C (HiCExplorer), DpnII, "
         "rep1+rep2 pooled per condition.", False),
        ("", False),
        ("Sheet 'compartment_calls': the 5-condition run (DN, DP, EbKO, ProB, "
         "3T3). 'call' = A if PC1 > 0, B if PC1 < 0. Cells shaded green (A) / "
         "red (B); bold marks a sign change relative to DN.", False),
        ("Sheet 'lymphoid_only_test': the 4-condition run (DN, DP, EbKO, ProB; "
         "3T3 excluded). This isolates the lymphoid comparison so the padj is "
         "not dominated by the fibroblast.", False),
        ("", False),
        ("IMPORTANT - how to read padj: dcHiC computes ONE adjusted p per bin "
         "that tests variation across ALL conditions in the run (a global, "
         "panel-level test), NOT a pairwise DN-vs-X difference. In the 5-cond "
         "run the padj is dominated by 3T3, so nearly every Tcrb bin is "
         "'global-significant' regardless of contrast. The reliable pairwise "
         "quantity is the compartment CALL / sign change read from the two PC1 "
         "columns. The lymphoid-only run gives the S3T3-free significance.",
         False),
        ("", False),
        ("Result summary: Tcrb is in the A compartment in all lymphoid "
         "conditions and B only in 3T3. In the lymphoid-only run, DP and EbKO "
         "show no bin crossing A->B; Pro-B crosses A->B at recombination-"
         "center-proximal bins (chr6:41.3-41.5 and 41.7-41.8 Mb; padj 1e-14 "
         "to 1e-20), while the domain core (41.0-41.3, 41.5-41.6 Mb) stays A.",
         False),
        ("", False),
        ("Provenance: generated by build_compartment_supp_table.py from the "
         "Tcrb_compartment_dchic_*.tsv and Tcrb_compartment_lymphoidonly_*.tsv "
         "region extracts (dchic_region_result.py). Run 2026-07-28.", False),
    ]
    for i, (txt, bold) in enumerate(notes, 1):
        c = ws3.cell(row=i, column=1, value=txt)
        c.font = Font(name=ARIAL, size=10, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws3.column_dimensions["A"].width = 110

    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    wb.save(a.out)
    print("wrote %s" % a.out)


if __name__ == "__main__":
    main()
